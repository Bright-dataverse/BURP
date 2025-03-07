# -*- coding: utf-8 -*-

__version__ = '1.9'
__author__ = 'Ir. V.Kroeze'

'''
BURP - Biogas Upgrading Report Program

Creates standardized reportsfor customers.
'''

#%% Import required libraries

import os
import math
from datetime import date

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

import pandas as pd
import numpy as np

#%% This is where the programm starts

class MonthlyReportingTool(tk.Tk):
    '''
    A GUI Tool for creating monthly reports based on standerdized datasets.
    
    This tool allows a user to create a report from IXON data without using a python interface. 
    When more installations are added the base code needs to be adjusted to allow for more sites to be added. 
    '''
    
    def __init__(self):
        super().__init__()
        '''
        This section allows for the inheritence of methods and the globalisation of variables within the instance
        '''
        
        self.title(f'Monthly Reporting Tool V{__version__}')
        
        '''Define some variables that are used in the tool'''
        self.site_var = tk.StringVar()
        self.file_location_var = tk.StringVar(value = '')
        self.folder_location_var = tk.StringVar(value = '')
        
        '''Start button is created glabally to allow for updating of state, initial state is disabeld'''
        self.start_button = ttk.Button(self, text='Start Reporting', command=self.start_reporting)
        
        '''Create the GUI widgets'''
        self.create_widgets()
        
        '''Check initial state of the start reporting button'''
        self.update_start_button_state()
    
    '''Methods of the reporting class are declared here'''
        
    def create_widgets(self):
        '''
        Create variables to configure the output format
        '''
        self.language_EN = tk.IntVar(value=0)

        self.Standard_B = tk.IntVar(value=0)
        self.Availability_B = tk.IntVar(value=0)
        self.Energy_B = tk.IntVar(value=0)

        self.Standard_C = tk.IntVar(value=0)
        self.Availability_C = tk.IntVar(value=0)
        self.Energy_C = tk.IntVar(value=0)

        self.Slip_BvsC = tk.IntVar(value=0)
        self.Energy_BnC = tk.IntVar(value=0)
        self.Heatpump = tk.IntVar(value=0)
        self.Energy_BnC_HP = tk.IntVar(value=0)
        '''
        Instantiates the lay out of the GUI
        '''
        
        label_site = ttk.Label(self, text="Select the installation:", font=("Calibri", 10))
        label_site.grid(row = 1, column = 1, padx = 5, pady = 25)
        
        site_select = ttk.Combobox(self, width=30, textvariable=self.site_var)
        site_select['values'] = (
            'B0175 - Aquafin NV',
            'B0218 - Delfland Harnaschpolder',
            'B0565 - Delfland Houtrust',
            'B0933 - Dommel',
            'H4187 - Twence',
            'H4242 - Delfland De Groote Lucht',
            'PR000041 - Dieckmann'
            )
        #site_select['state'] = 'readonly'
        site_select.current()
        site_select.grid(row = 1, column = 2, padx = 5, pady = 25)
        
        label_language = ttk.Label(self, text="Select Box for English:", font=("Calibri", 10))
        label_language.grid(row = 2, column = 1, padx = 5, pady = 25)

        language_EN_box = ttk.Checkbutton(self, text='', variable=self.language_EN)
        language_EN_box.grid(row = 2, column = 2, padx = 5, pady = 25)

        label_B_Options = ttk.Label(self, text="Select Export options Upgrader:", font=("Calibri", 10))
        label_B_Options.grid(row = 3, column = 1, padx = 5, pady = 25)

        B_Standard_box = ttk.Checkbutton(self, text='Standard', variable=self.Standard_B)
        B_Standard_box.grid(row = 3, column = 2, padx = 5, pady = 25)

        B_Availability_box = ttk.Checkbutton(self, text='Availability', variable=self.Availability_B)
        B_Availability_box.grid(row = 3, column = 3, padx = 5, pady = 25)

        B_Energy_box = ttk.Checkbutton(self, text='Energy', variable=self.Energy_B)
        B_Energy_box.grid(row = 3, column = 4, padx = 5, pady = 25)

        label_C_Options = ttk.Label(self, text="Select Export options Liquefaction:", font=("Calibri", 10))
        label_C_Options.grid(row = 4, column = 1, padx = 5, pady = 25)

        C_Standard_box = ttk.Checkbutton(self, text='Standard', variable=self.Standard_C)
        C_Standard_box.grid(row = 4, column = 2, padx = 5, pady = 25)

        C_Availability_box = ttk.Checkbutton(self, text='Availability', variable=self.Availability_C)
        C_Availability_box.grid(row = 4, column = 3, padx = 5, pady = 25)

        C_Energy_box = ttk.Checkbutton(self, text='Energy', variable=self.Energy_C)
        C_Energy_box.grid(row = 4, column = 4, padx = 5, pady = 25)

        label_BnC_Options = ttk.Label(self, text="Select Export options Combined installations:", font=("Calibri", 10))
        label_BnC_Options.grid(row = 5, column = 1, padx = 5, pady = 25)

        Slip_BvsC_box = ttk.Checkbutton(self, text='Slip Biogas vs Liquefaction', variable=self.Slip_BvsC)
        Slip_BvsC_box.grid(row = 5, column = 2, padx = 5, pady = 25)

        BnC_Energy_box = ttk.Checkbutton(self, text='Energy', variable=self.Energy_BnC)
        BnC_Energy_box.grid(row = 5, column = 3, padx = 5, pady = 25)

        Heatpump_box = ttk.Checkbutton(self, text='Heatpump', variable=self.Heatpump)
        Heatpump_box.grid(row = 5, column = 4, padx = 5, pady = 25)

        Energy_BnC_HP_box = ttk.Checkbutton(self, text='Energy Heatpump', variable=self.Energy_BnC_HP)
        Energy_BnC_HP_box.grid(row = 5, column = 8, padx = 5, pady = 25)

        select_file_button = ttk.Button(self, text='Data File', command=self.select_file)
        select_file_button.grid(row = 6, column = 1, padx = 5, pady = 25)
        
        show_open_file = ttk.Label(self, textvariable=self.file_location_var, font=("Calibri", 10))
        show_open_file.grid(row = 6, column = 2, padx = 5, pady = 25)
        
        select_folder_button = ttk.Button(self, text='Destination folder', command=self.select_folder)
        select_folder_button.grid(row = 7, column=1, padx = 5, pady = 25)
        
        show_open_folder = ttk.Label(self, textvariable=self.folder_location_var, font=("Calibri", 10))
        show_open_folder.grid(row = 7, column=2, padx = 5, pady = 25)
        
        '''Starts the generation of the report, initially disabled'''
        self.start_button.grid(row = 8, column=1, padx = 5, pady=25)
        self.start_button.config(state=tk.DISABLED)
        
        '''Resets the input fields to empty'''
        reset_button = ttk.Button(self, text="Reset", command=self.reset)
        reset_button.grid(row=8, column=2, pady=25)
        
        '''stops the program'''
        cancel_button = ttk.Button(self, text="Cancel", command=self.destroy)
        cancel_button.grid(row=8, column=3, padx = 25, pady=25)
        
        '''Updates the state of variables to enable start button'''
        self.site_var.trace("w", lambda *args: self.update_start_button_state())
        self.file_location_var.trace("w", lambda *args: self.update_start_button_state())
        self.folder_location_var.trace("w", lambda *args: self.update_start_button_state())
        
    def select_file(self):
        filetypes = [('csv files', '*.csv'), ('All files', '*.*')]
        filename = filedialog.askopenfilename(title='Open: Monthlyreport_data.csv', initialdir='./user/Downloads', filetypes=filetypes)
        self.file_location_var.set(filename)
        
    def select_folder(self):
        foldername = filedialog.askdirectory()
        self.folder_location_var.set(foldername)
        
    def update_start_button_state(self):
        '''
        When opening the GUI, initially the start button will be disabled
        It is enabled when all variable field contain a value

        '''
        site_selected = bool(self.site_var.get())
        file_selected = bool(self.file_location_var.get())
        folder_selected = bool(self.folder_location_var.get())
        self.start_button.config(state=tk.NORMAL if site_selected and file_selected and folder_selected else tk.DISABLED)
    
    def reset(self):
        print(self.language_EN.get())
        self.site_var.set('')
        self.file_location_var.set('')
        self.folder_location_var.set('')
        pass
    
    '''
    Here we generate the different reports
    
    When more installations are added, this section needs to be increased
    '''
    
    def start_reporting(self):
        '''Here we initialize the process of reporting'''
        
        report_generator = StandardizedReport(self.site_var.get(), self.file_location_var.get(), self.folder_location_var.get())
        
        report_export = ExportToExcel(report_generator,self.language_EN.get(),self.Standard_B.get(),self.Availability_B.get(),self.Energy_B.get(),self.Standard_C.get(),self.Availability_C.get(),self.Energy_C.get(),self.Slip_BvsC.get(),self.Energy_BnC.get(),self.Heatpump.get(),self.Energy_BnC_HP.get())

        pass

class StandardizedReport:
    '''
    Class to represent the information that has to go into the report.
    Recipes are defined for every installation.
    Elements of these receipes are defined below.
    '''
    
    def __init__(self, site, file_location, folder_location):
        self.site = site
        self.file_location = file_location
        self.folder_location = folder_location
        
        ''' The following elements are general across al reports'''
        self.monthly_report_database = self.import_csv()
        
        self.error_list = self.create_error_list(self.monthly_report_database, 'SEQSTATE')
        
        self.monthly_report_database = self.resample_data(self.monthly_report_database)
        
        self.calculate_basic_data(self.monthly_report_database)
        
        self.installation_specific_data(self.monthly_report_database)
        
        
        
    def import_csv(self):
        '''First import the csv with the data'''
        df = pd.read_csv(self.file_location)
        
        '''Next do the first transformations'''
        
        df['time'] = pd.to_datetime(df['time'])
        df.sort_values(by=['time'], inplace = True)
        df.reset_index(drop=True, inplace = True)
        
        df.fillna(method = 'ffill', inplace = True)
        df.fillna(method = 'bfill', inplace = True)
        
        # Some time data for saving purposes
        self.period = df['time'].dt.strftime('%B-%Y')[0]
        self.save_period = df['time'].dt.strftime('%Y-%m')[0]
        
        return df
    
    def create_error_list(self, dataframe, tag):
        
        df = dataframe
        tag = tag
        Error = False # priming the check variable
        time_start = [] # priming a time counter
        time_duration = []
        time_end = []
        Error_data = []
        i = 0

        # Iterate over the dataframe
        for index, row in df.iterrows():
                
            if (row[tag] == 90 or row[tag] == 99) and Error == False:
                Error = True
                time_start = row['time']
                Error_data.append(row['time'])
                time_duration.append(row['time'] - time_start)
                time_end.append(row['time'])
            elif (row[tag] == 90 or row[tag] == 99):
                Error = True
            elif (row[tag] != 90 or row[tag] != 99) and Error == True:
                time_duration[i] = (row['time'] - time_start)
                time_end[i] = (row['time'])
                i = i+1
                Error = False
            else:
                Error = False
        
        trip_data = pd.DataFrame([Error_data, time_duration, time_end]).T
        trip_data.columns = ['Date','Duration', 'endDate']
        
        return trip_data
    
    def resample_data(self, dataframe):
        
        df = dataframe
        df = df.set_index('time').resample('300S').bfill()
        
        return df
    
    def calculate_basic_data(self, dataframe):
        
        df = dataframe

        '''Get the basic data for the report'''

        self.biogas = df[df['SEQSTATE']==62]['RHA10CF001'].sum()/12 # from 5 minute data, sum of biogas flow in production [Nm3]
        self.biogas_CH4 = df[(df['SEQSTATE']==62) & (df['RHH15_CH4']>25)]['RHH15_CH4'].mean() # from 5 minute data, average of biogas methane [%]

        self.biomethane = df[df['SEQSTATE']==62]['NormalFlow'].sum()/12 # from 5 minute data, sum of biomethane flow in production [Nm3]
        self.biomethane_CH4 = df[df['SEQSTATE']==62]['RHH10_CH4'].mean() # from 5 minute data, average of biomethane methane [%]

        self.capacity = (50+(df[df['SEQSTATE']==62]['RHM50AN001']/2)-(df[df['SEQSTATE']==62]['RHM50AA106']/2)).mean()
        self.methane_slip = 100*(1-(self.biomethane*self.biomethane_CH4/self.biogas/self.biogas_CH4)) # Guestimate for slip, must be measured [%]

        ''''Availability hours'''

        self.trip = math.floor(df[(df['SEQSTATE'] == 90) | (df['SEQSTATE']==99)]['SEQSTATE'].count()/12)
        self.standby = math.floor(df[df['SEQSTATE'] == 1]['SEQSTATE'].count()/12)
        self.running = math.ceil((df['SEQSTATE'].count()/12)-self.trip-self.standby)
    
    def installation_specific_data(self, dataframe):
        
        df = dataframe
        
        if self.site == 'B0565 - Delfland Houtrust':

            (self.trip_CO2LIQ, self.standby_CO2LIQ, self.running_CO2LIQ) = self.calculate_availability(self.monthly_report_database, 'CO2LIQ')
            (self.trip_heatpump, self.standby_heatpump, self.running_heatpump) = self.calculate_availability(self.monthly_report_database, 'Heatpump')
            self.total_energy = self.calculate_energy(self.monthly_report_database, 'Energy')
            self.total_energy_CO2 = self.calculate_energy(self.monthly_report_database, 'Energy_CO2')
            self.total_energy_HP = self.calculate_energy(self.monthly_report_database, 'Energy_HP')
            
        elif self.site == 'B0933 - Dommel':
            self.total_energy = self.calculate_energy(self.monthly_report_database, 'Energy')
            (self.trip_CO2LIQ, self.standby_CO2LIQ, self.running_CO2LIQ) = self.calculate_availability(self.monthly_report_database, 'SEQSTATE_CO2', normal = False)
            self.methane_slip = (df[(df['SEQSTATE']==62) & (df['SEQSTATE_CO2'].isin([1,2,90,99]))]['Methane_slip'].mean()/100) * \
                (df[(df['SEQSTATE']==62) & (df['RHH15_CH4']>25)]['RHH15_CH4'].mean()/100) * \
                (101325*16.04/8.314/273.15)
            self.methane_slip_CO2LIQ_active = (df[(df['SEQSTATE']==62) & (df['SEQSTATE_CO2']==20)]['Methane_slip'] * \
                (df[(df['SEQSTATE']==62) & (df['SEQSTATE_CO2']==20)]['Methane_slip_factor']/100)).mean()
            self.total_energy_CO2 = self.calculate_energy(self.monthly_report_database, 'Energy_CO2_2 (kWh)')
            self.mean_H2S_treated = self.calculate_H2S_correction_Dommel(self.monthly_report_database)
            self.error_list_CO2 = self.create_error_list(self.monthly_report_database, 'SEQSTATE_CO2')
            
        elif self.site == 'H4187 - Twence':
            self.methane_slip = 'n.v.t.'
            
        elif self.site == 'PR000041 - Dieckmann':
            (self.trip_CO2LIQ, self.standby_CO2LIQ, self.running_CO2LIQ) = self.calculate_availability(self.monthly_report_database, 'SEQSTATE_CO2', normal = False)
            
        else:
            pass
            
    def calculate_availability(self, dataframe, column_name, normal = True):
        
        df = dataframe
        
        '''Availability hours'''
        
        if normal == True:
            trip = math.floor(df[df[column_name] == 5][column_name].count()/12)
            standby = math.floor(df[df[column_name] == 1][column_name].count()/12)
            running = math.ceil((df[column_name].count()/12)-trip-standby)
        else:
            trip = math.floor(df[(df[column_name] == 90) | (df[column_name] == 99) | (df[column_name] == 1)][column_name].count()/12)
            standby = math.floor(df[df[column_name] == 2][column_name].count()/12)
            running = math.ceil((df[column_name].count()/12)-trip-standby)
        
        return trip, standby, running

    def calculate_energy(self, dataframe, column_name):
        
        df = dataframe
        
        total_energy = 0
        previous = None

        for index, row in df.iterrows():
            current = row[column_name]
            
            if previous is None:
                previous = current
                continue
            
            energy_diff = current-previous
            
            if energy_diff < 0:
                energy_diff = 0
            
            total_energy += energy_diff
            previous = current
        return total_energy
    
    def calculate_H2S_correction_Dommel(self,dataframe):
        df = dataframe
        
        df['H2S_in'] = df['H2S_in'].apply(lambda val: np.nan if val < 5 else val)
        df['H2S_in'].interpolate(method='linear', inplace=True)
        df = df[df['SEQSTATE'] == 62]
        mean_H2S = (df['H2S_in']*df['RHA10CF001']).sum()/df['RHA10CF001'].sum()
        
        return mean_H2S


class ExportToExcel:
    
    def __init__(self, StandardizedReport, language_EN, B_Standard, B_Availability, B_Energy, C_Standard, C_Availability, C_Energy, Slip, Energy_BnC, Heatpump, Energy_BnC_HP):
        
        self.exd = StandardizedReport # exd = external data, shortened for ease of use
        
        self.Language_EN = language_EN

        available_templates = {"B_Standard": B_Standard,"C_Standard":C_Standard, "B_Availability": B_Availability,"C_Availability": C_Availability, "Slip": Slip,"B_Energy":B_Energy,"C_Energy": C_Energy,"BnC_Energy":Energy_BnC,"Heatpump":Heatpump,"EnergyHeatpump":Energy_BnC_HP}
        selected_templates = [os.getcwd()+f"/BaseFiles/{key}_{language_EN}.xlsx" for key, value in available_templates.items() if value == 1]
        self.template_files = selected_templates
        self.build_final_excel()

    def copy_block(self,source_ws, target_ws, start_row, start_col):
        """Copy a block of data from source worksheet to target worksheet at a specific row and column."""
        max_row = source_ws.max_row
        max_col = source_ws.max_column
    
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                source_cell = source_ws.cell(row=row, column=col)
                target_cell = target_ws.cell(row=start_row + row - 1, column=start_col + col - 1)
                
                # Copy value
                target_cell.value = source_cell.value
                
                # Copy style attributes safely
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)
        
        return max_row, max_col  # Return dimensions of the copied block

    def build_final_excel(self):
        """Builds the final Excel file by arranging blocks in a grid format."""
        nvt = lambda x: x if isinstance(x, str) else round(x, 2)

        final_wb = load_workbook(filename =  os.getcwd()+f'/BaseFiles/0Header_{self.Language_EN}.xlsx')
        final_ws = final_wb.active

        final_ws['C4'] = self.exd.site
        final_ws['C5'] = self.exd.period
        final_ws['C6'] = f'{date.today()}'
                
        current_row = 10
        current_col = 2
        max_row_in_grid = 10  # Keeps track of row height in the current grid row
        
        for template in self.template_files:
            wb = load_workbook(template)
            ws = wb.active  # Assume data is in the first sheet

            # Fill the partial template with the appropriate data
            if "B_Standard" in template:
                ws['B2'] = round(self.exd.biogas,1)
                ws['B3'] = round(self.exd.biogas_CH4,1)

                ws['B6'] = round(self.exd.biomethane,1)
                ws['B7'] = round(self.exd.biomethane_CH4,1)

                ws['B9'] = round(self.exd.capacity)
                ws['B10'] = nvt(self.exd.methane_slip)

            if "B_Availability" in template:
                ws['B2'] = self.exd.running
                ws['B3'] = self.exd.standby
                ws['B4'] = '='+str(self.exd.trip)+'-G14-G15-G16'
                ws['B5'] = 0
                ws['B6'] = 0
                ws['B7'] = 0
                ws['B8'] = '=100*(1-((G13+G14+G15+G16)/(G11+G12+G13+G14+G15+G16)))'

            if "C_Availability" in template:
                ws['B2'] = self.exd.running_CO2LIQ
                ws['B3'] = self.exd.standby_CO2LIQ
                ws['B4'] = '='+str(self.exd.trip_CO2LIQ)+'-G18-G19-G20'
                ws['B5'] = 0
                ws['B6'] = 0
                ws['B7'] = 0
                ws['B8'] = '=100*(1-((G17+G18+G19+G20)/(G15+G16+G17+G18+G19+G20)))'

            if "Heatpump" in template:
                ws['B2'] = self.exd.running_heatpump
                ws['B3'] = self.exd.standby_heatpump
                ws['B4'] = self.exd.trip_heatpump
                ws['B5'] = '=100*(1-(G24/(G22+G23+G24)))'

            if "EnergyHeatpump" in template:
                ws['B2'] = self.exd.total_energy
                ws['B3'] = self.exd.total_energy_HP
                ws['B4'] = self.exd.total_energy_CO2
                
                ws['F2'] = 1000*self.exd.total_energy/self.exd.biogas
                ws['F3'] = 1000*self.exd.total_energy_HP/self.exd.biogas
                ws['F4'] = 1000*self.exd.total_energy_CO2/self.exd.biogas

            if "Slip" in template:
                ws['B2'] = round(self.exd.methane_slip,2)
                ws['F2'] = self.exd.methane_slip_CO2LIQ_active

            if "B_Energy" in template:
                ws['B2'] = self.exd.total_energy
                ws['B3'] = 1000*self.exd.total_energy/self.exd.biogas

            if "C_Energy" in template:	
                ws['B2'] = self.exd.total_energy_CO2/1000000
                ws['B3'] = self.exd.total_energy_CO2/self.exd.biogas/1000
                        
            # Copy the block to the final worksheet    

            block_rows, block_cols = ws.max_row, ws.max_column
            
            # Check if the block fits in the current row
            if current_col + block_cols > 9:  # 1 + 7 columns + 1 for spacing
                current_row = final_ws.max_row + 2  # Move to next grid row
                current_col = 2
                max_row_in_grid = 10
            self.copy_block(ws, final_ws, current_row, current_col)
            max_row_in_grid = max(max_row_in_grid, block_rows)
            
            # Update column position for next block
            current_col += block_cols + 1  # Add spacing column

        wb1 = load_workbook(filename =  os.getcwd()+f'/BaseFiles/1_Triplist_{self.Language_EN}.xlsx')
        ws1 = wb1.active  # Assume data is in the first sheet

        current_col = 2
        current_row = final_ws.max_row +2
        self.copy_block(ws1, final_ws, current_row, current_col)

        skip = final_ws.max_row +1  

        for i in range(skip,skip+len(self.exd.error_list)):
            final_ws.insert_rows(idx = i)
            ws.merge_cells(start_row=i, start_column=4, end_row=i, end_column=6)
            thin = Side(border_style="thin", color="000000")
            final_ws.cell(row = i, column = 2).border = Border(left = thin)
            final_ws.cell(row = i, column = 8).border = Border(right = thin)
            final_ws.cell(row = i, column = 2, value = self.exd.error_list['Date'][i-skip].strftime('%Y-%m-%d %H:%M'))
            final_ws.cell(row = i, column = 3, value = math.ceil(100*self.exd.error_list['Duration'][i-skip].total_seconds()/60/60)/100)
            final_ws.cell(row = i, column = 7, value = self.exd.error_list['endDate'][i-skip].strftime('%Y-%m-%d'))

        max_row_in_grid = skip+len(self.exd.error_list)

        wb2 = load_workbook(filename =  os.getcwd()+f'/BaseFiles/2_Closing_{self.Language_EN}.xlsx')
        ws2 = wb2.active  # Assume data is in the first sheet

        current_col = 2
        current_row = final_ws.max_row +2
        self.copy_block(ws2, final_ws, current_row, current_col)

        final_wb.save(filename = self.exd.folder_location+'/'+self.exd.site+' - maandrapportage bedrijfsvoering '+self.exd.save_period+'.xlsx')
        
        messagebox.showinfo('Information', 'Report for '+self.exd.period+' '+self.exd.site+' was created')

        '''
        ws['C22'] = self.exd.running_CO2LIQ
        ws['C23'] = self.exd.standby_CO2LIQ
        ws['C24'] = self.exd.trip_CO2LIQ
        ws['C26'] = '=100*(1-(C24/(C22+C23+C24)))'

        template 2 short CO2 availability
        '''
'''This starts up an instance of our application'''

def main():
    tool = MonthlyReportingTool()
    tool.mainloop()

if __name__ =='__main__':
    main()
    
    
    